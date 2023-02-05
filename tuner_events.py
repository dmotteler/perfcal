#!/usr/bin/env python
'''
    started with
    $Id: tuner_events.py,v 1.12 2022/04/04 03:10:30 dfm Exp dfm $

    removed database parts, using openpyxl for spreadsheet access

'''

import os, sys
from datetime import datetime, timedelta, timezone
from icalendar import Calendar, Event, tools
import pytz
from calendar import month_name
from zipfile import ZipFile
from openpyxl import load_workbook
import msoffcrypto
import io
from copy import copy

class tuner_events():

    def __init__(self, infile, dotypes, caln, outext=None, fromdate=None, todate=None):
        self.infile = infile
        self.outext = outext
        self.calnames = caln

        self.pst = pytz.timezone("US/Pacific")

        self.fromdate = fromdate.astimezone(self.pst)
        self.todate = todate.astimezone(self.pst)

        self.dotypes = dotypes

        self.venue_addrs = {}

        self.events = {}

        if self.infile is None:
            self.event_source = "none"
            self.event_class = self
        elif ".xls" in self.infile:
            self.exc_events()
            self.event_source = "xls"
            self.event_class = self
        else:
            self.ics_events()
            self.event_source = "ical"
            self.event_class = self

    def overlap(self, evstrt, evend, title):
        ''' return 0 => no overlap, 1 => start and end match a previous event, 2 => some overlap '''

        s_e = (evstrt, evend)
        rv = 0

        if s_e in self.events:
            rv = 1
            (evst, evnd) = s_e
        else:
            # no event with matching start and end times. Still need
            # to see if the new event overlaps one we already knew about.

            for s_e in self.events:
                (evst, evnd) = s_e
                if evend <= evst or evstrt >= evnd:
                    continue
                rv = 2
                break

        if rv > 0:
            sdat1 = evst.strftime("%m/%d/%Y")
            timstr1 = evst.strftime("%H:%M %p")
            timend1 = evnd.strftime("%H:%M %p")
            sdat2 = evstrt.strftime("%m/%d/%Y")
            timstr2 = evstrt.strftime("%H:%M %p")
            timend2 = evend.strftime("%H:%M %p")
            print("event overlap:    date      start    end   event")
            print("               %s %s %s %s" % (sdat1, timstr1, timend1, self.events[s_e]['title']))
            print("               %s %s %s %s" % (sdat2, timstr2, timend2, title))
            if rv == 1:
                print("....SECOND EVENT DISCARDED....")
            print()

        return rv

    def exc_events(self):
        ''' load events in date range from excel spreadsheet '''

        # openpyxl doesn't support data validation, but we don't care. suppress the warning.
        import warnings
        warnings.filterwarnings("ignore", "Data Validation")

        # the SingoutInfo workbook may be password protected to shelter from
        # prying eyes. if SIPW is in the environment, decrypt the file
        # before opening.

        sipw = os.environ.get('SIPW', None)
        if sipw:
            # from stackoverflow.com/questions/19450837/how-to-open-a-password-protected-excel-file-using-python
            decrypted_workbook = io.BytesIO()
            with open(self.infile, 'rb') as file:
                office_file = msoffcrypto.OfficeFile(file)
                office_file.load_key(password=sipw)
                office_file.decrypt(decrypted_workbook)

            self.wb = load_workbook(decrypted_workbook, read_only=True)
        else:
            self.wb = load_workbook(self.infile, read_only=True)

        sh = self.wb["venues"]
        numcols = 3

        for row in sh.iter_rows(min_row=2, values_only=True):
            ven, add1, add2 = list(row)[:3]
            if ven is None:
                break
            self.venue_addrs[ven] = (add1, add2)

        # print(self.dotypes)

        yrs = sorted([int(self.fromdate.year), int(self.todate.year) + 1])
        for yr in range(yrs[0], yrs[1]):
            if self.dotypes['p']:
                self.dosheet("Performances", yr)

            if self.dotypes['r']:
                self.dosheet("Rehearsals", yr)

            if self.dotypes['b']:
                self.dosheet("board mtgs", yr)

            if self.dotypes['a']:
                self.dosheet("absences", yr)

        ## ? self.wb.Close(False)

    def dosheet(self, evtypes, yr):
        # evtypes is "Performances" or "Rehearsals" or "board mtgs" or "absences"
        perfsheet = "%d %s" % (yr, evtypes)
        # print("dosheet {}".format(perfsheet))

        try:
            sh = self.wb[perfsheet]
        except Exception:
            print("no such sheet: {}".format(perfsheet))
            return

        hdrs = list(list(sh.iter_rows(max_row=1, values_only=True))[0])

        # remove None hdrs from end of list
        ndx = len(hdrs) - 1
        while ndx >= 0 and hdrs[ndx] is None:
            hdrs.pop()
            ndx = len(hdrs) - 1

        # print("sheet %s has %d columns" % (perfsheet, len(hdrs)))
        # print("fromdate: %s, todate: %s" % (self.fromdate, self.todate))

        if evtypes == "absences":
            for row in sh.iter_rows(min_row=2, values_only=True):
                stdate, enddate, desc = list(row)

                if stdate is None:
                    break # empty date column signals end of data
                if desc is None:
                    continue # probably should complain?

                evstart = self.mdydate(stdate, 0, 0, 0)
                if enddate is None:
                    enddate = stdate

                evend = self.mdydate(enddate, 23, 59, 59)
                # evend = self.mdydate(enddate, 0, 0, 0)

                if evend < evstart:
                    print("Ignoring %s absence entry: end(%s) is before start(%s)!" % (desc, evend, evstart))
                    continue

                if evend < self.fromdate or evstart > self.todate:
                    # event outside requested range - skip it.
                    continue

                # google calendar has a bug - doesn't include the 
                # last day of a multi-day event.
                evend += timedelta(days=1)

                s_e = (evstart, evend)
                self.events[s_e] = {}

                self.events[s_e]['title'] = desc
                self.events[s_e]['type'] = evtypes
                # set venue to blank for listing
                self.events[s_e]['venue'] = ""

            return

        for row in sh.iter_rows(min_row=2, values_only=True):
            evtitl, venue, evdate, sttime, endtime, uni, evtype = list(row[:7])

            if evdate is None:
                break # empty date column signals end of data

            if evtype != "absences" and venue not in self.venue_addrs:
                print("didn't recognize %s as a valid venue???" % (venue))
                continue

            if evtitl is None:
                continue # date but no title implies not booked - just skip it.

            evdate = self.dtdate(evdate)
            evstart = self.sethm(evdate, sttime)
            evend = self.sethm(evdate, endtime)

            overl = self.overlap(evstart, evend, evtitl)
            if overl == 1:
                continue

            if evend < self.fromdate or evstart > self.todate:
                # event outside requested range - skip it.
                continue

            s_e = (evstart, evend)
            if s_e in self.events:
                timstr = evstart.strftime("%m/%d/%Y at %H:%M %p")
                print("duplicate event start/end date/time: event1: %s, event2: %s, both on %s" % (self.events[s_e]['title'], evtitl, timstr))
                continue

            self.events[s_e] = {}

            self.events[s_e]['title'] = evtitl
            self.events[s_e]['venue'] = venue
            ## self.events[s_e]['evend'] = evend
            self.events[s_e]['uni'] = uni
            self.events[s_e]['type'] = evtype

    def mdydate(self, d, h, m, s):
        # receive a datetime.datetime (excel date) from the absences spreadsheet,
        # and the values for hours, minutes, and seconds to init and
        # return a datetime object

        try:
            t = d.timetuple()
            dout = self.pst.localize(datetime(t.tm_year, t.tm_mon, t.tm_mday, h, m, s))
        except Exception as e:
            print("%s: %s" % (e, str(d)))
            dout = None

        print("mdydate returns %r for %s %d %d %d" % (dout, d, h, m, s))
        return dout

    def dtdate(self, d):
        # receive date like yyyy-mm-dd HH:MM:SS
        # return datetime object inited to mm/dd/yy HH:MM:SS PST

        sdate = str(d)
        try:
            dout = self.pst.localize(datetime.strptime(sdate[:19], "%Y-%m-%d %H:%M:%S"))
        except Exception as e:
            print("%s: %s" % (e, sdate))
            dout = None

        # print("dtdate returning {} for {}".format(dout, d))
        return dout

    def sethm(self, dt, t):
        # given a datetime dt and time t,
        # return dt with hours and minutes replaced

        dout = copy(dt)
        if t:
            dout = dout.replace(hour=t.hour, minute=t.minute)

        # print("dt {}, t {}, out {}".format(dt, t, dout))
        return dout

    def list_events(self):
        print("\nListing events from %s" % (self.infile))

        for s_e in sorted(self.events):
            (evst, evend) = s_e
            ev = self.events[s_e]
            if 'type' in ev:
                typ = ev['type']
            else:
                typ = ""

            if typ == "absences" or evst.hour == 0:
                # we'll assume that no event really started at midnight,
                # so this is from absences
                sts = evst.strftime("%m/%d/%Y")
                ends = evend.strftime("%m/%d/%Y")

                print("%s"  % (ev['title']))
                print("  from %s to %s"  % (sts, ends))
                print("  type: %s\n"  % (typ))
            else:
                ven = ev['venue']
                if ven in self.venue_addrs:
                    a1, a2 = self.venue_addrs[ven]
                else:
                    print("Venue %s unknown by %s type" % (ven, self.event_source))
                    a1 = a2 = "??"

                sts = evst.strftime("%m/%d/%Y at %I:%M %p")
                ends = evend.strftime("%m/%d/%Y at %I:%M %p")

                print("%s at %s"  % (ev['title'], ven))
                print("  %s\n  %s" % (a1, a2))
                print("  from %s to %s"  % (sts, ends))
                print("  uni: %s, type: %s\n"  % (ev['uni'], typ))

    def do_cal(self, mem, caldata):
        ''' add events from ics calendar to list of events '''

        cal = Calendar()
        gcal = cal.from_ical(caldata)
        nev = 0

        for sub in gcal.subcomponents:
            if sub.name == "VEVENT":
                uid = sub['UID']
                # print("uid from ics %s" % (uid))

                evstrt = sub['DTSTART'].from_ical(sub['DTSTART'])
                if not isinstance(evstrt, datetime):
                    # absences (and others?) s/b "all day", giving a "date", not "datetime"
                    evstrt = self.pst.localize(datetime(evstrt.year, evstrt.month, evstrt.day))
                else:
                    evstrt = evstrt.astimezone(self.pst)
                    # it's a datetime - if we're loading an absences calendar, inputs from
                    # other than .ics have midnite today, as today 0:0:0
                    if "abs" in mem:
                        if evstrt.hour != 0:
                            bump = evstrt.hour
                            print("subtracting {} hours from {}".format(bump, evstrt))
                            evstrt -= timedelta(hours=bump)

                evend = sub['DTEND'].from_ical(sub['DTEND'])
                if not isinstance(evend, datetime):
                    # better be a "date" if it's not a "datetime". turn it into a datetime
                    evend = self.pst.localize(datetime(evend.year, evend.month, evend.day))
                else:
                    evend = evend.astimezone(self.pst)
                    # it's a datetime - if we're loading an absences calendar, inputs from
                    # other than .ics have midnite tonite, as tomorrow 0:0:0
                    if "abs" in mem:
                        if evend.hour != 0:
                            bump = 24 - evend.hour
                            print("adding {} hours to {}".format(bump, evend))
                            evend += timedelta(hours=bump)

                if evend < self.fromdate or evstrt > self.todate:
                    # event outside requested range - skip it.
                    continue

                uni = ""
                typ = ""

                if 'DESCRIPTION' in sub:
                    descs = sub['DESCRIPTION'].split("\n")
                    for desc in descs:
                        if desc.startswith("UNIFORM:"):
                            uni = desc[len("UNIFORM:"):]
                        elif desc.startswith("EVENT_TYPE:"):
                            typ = desc[len("EVENT_TYPE:"):]

                venue = ""
                loc = sub['LOCATION']
                if loc != "":
                    locflds = loc.split("\n")
                    venue = locflds.pop(0).strip()
                    if len(locflds) > 0:
                        loc = "\n".join(locflds)
                    else:
                        locflds = [""]
                    self.venue_addrs[venue] = locflds

                title = sub['SUMMARY']

                s_e = (evstrt, evend)

                if s_e in self.events:
                    print("conflict:")
                    print("  start, end times: %s, %s" % (evstrt.strftime("%Y-%m-%d %H:%M"),
                        evend.strftime("%Y-%m-%d %H:%M")))
                    print("  name 1: %s" % (self.events[s_e]['title']))
                    print("  name 2: %s" % (title))
                    
                else:
                    if typ == "Rehearsal" and not self.dotypes['r']:
                        continue
                    elif typ in ["Performance", "Other", "Social Event"] and not self.dotypes['p']:
                        continue
                    elif typ == "Meeting" and not self.dotypes['b']:
                        continue

                    self.events[s_e] = {'title': title, 'venue': venue, 'addr': loc, 'uni': uni, 'type': typ, 'uid': uid}
                    nev += 1

        self.calfiles[mem][1] = nev

    def ics_events(self):
        ''' process an ical-type file, create a list of events '''

        self.calfiles = {}
        oldcal = None

        self.events = {}

        if self.infile.endswith(".ics"):
            fo = open(self.infile, "r")
            oldcal = fo.read()
            fo.close()

            self.calfiles[self.infile] = [self.infile, 0]
            self.do_cal(self.infile, oldcal)

        elif self.infile.endswith(".zip"):
            fz = ZipFile(self.infile)

            yrs = sorted([int(self.fromdate.year), int(self.todate.year) + 1])
            knowncals = fz.namelist()

            if not self.calnames:
                self.calnames = []
                for yr in range(yrs[0], yrs[1]):
                    self.calnames.append("tuners{}_".format(yr))
                    if self.dotypes['a']:
                        self.calnames.append("tunersboardabs{}_".format(yr))

            for caln in self.calnames:
                found = False
                for mem in knowncals:
                    if mem.startswith(caln):
                        self.calfiles[caln] = [mem, 0]
                        found = True
                        break
                if not found:
                    print("no such calendar: {}".format(caln))

            for calname in self.calfiles:
                mem = self.calfiles[calname][0]
                oldcal = fz.read(mem)
                self.do_cal(calname, oldcal)

            fz.close()
